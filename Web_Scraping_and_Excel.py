import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from time import localtime, strftime

## Date Stamp ##
month = strftime("%m",localtime())
day = strftime("%d", localtime())
year = strftime("%Y", localtime())
datestamp = f"{month}_{day}_{year}"
#print(f"{month}_{day}_{year}")

## Selenium ##
#browser = webdriver.Chrome("C:\\Users\\gnjg1\\chromedriver.exe")
browser = webdriver.Chrome()

## Openpyxl ##

wkbk = Workbook()
#testWB = openpyxl.load_workbook("Test2.xlsx")
testWB = openpyxl.load_workbook(filename = "FCS Bank Leadership.xlsm",keep_vba = True)
sheets = testWB.worksheets
bankSheet = testWB.copy_worksheet(sheets[0])
bankSheet.title = datestamp
testWB.active = bankSheet
bankSheet["A1"] = datestamp
masterSheet = testWB.worksheets[1]
#bankSheet = testWB.active

## for the compare
#lastSheet = testWB[sheets[1]] 
#lastSheet = testWB.worksheets[-2]

def texasfcb():
    
    browser.get("https://www.farmcreditbank.com/about-us/senior-management/")
    #csuite = browser.find_elements_by_tag_name("div.col-17.pl-15")
    csuite = browser.find_elements_by_tag_name("div.col-24.offset-md-1.col-md-15.offset-lg-0.col-lg-16.col-xl-16.pl-0.pl-md-10.pl-lg-10.pl-xl-10.pt-15.pt-md-0.pt-lg-0.pt-xl-0")

    title = []
    name = []
    #form changes based on size...
    nameList = []
    prevLeadership = set()

    for x in range(3,20):
        while masterSheet[f"H{str(x)}"].value != None:
            prevLeadership.add(masterSheet[f"H{str(x)}"].value)
            break

    for item in csuite:
        nm_item = item.find_element_by_tag_name("h4.underline-green")
        ttl_item = item.find_element_by_tag_name("h6")

        title.append(ttl_item)
        name.append(nm_item)

    leadership = {}

    if len(name) == len(title):
        for i in range(0,len(name)):
            leadership[name[i].text] = title[i].text
            nameList.append(name[i].text)
            i += 1

        for i in range(0,len(name)):
            bankSheet["K"+ str(5+i)] = name[i].text
            bankSheet["L"+ str(5+i)] = leadership[name[i].text]
            bankSheet["M"+ str(5+i)] = datestamp

    setCompare = prevLeadership.symmetric_difference(nameList)
    
    if len(setCompare) > 0:
        bankSheet[f"K{5 + len(name)}"] = str(setCompare)
        bankSheet[f"K{5 + len(name)}"].fill = PatternFill("solid", fgColor="FFA7A7")      
        
        for x in range(2,20):
            while masterSheet[f"H{str(x)}"].value != None:
                masterSheet[f"H{str(x)}"].value = None
                break
        for i in range(0,len(name)):
            masterSheet["H"+ str(3+i)] = name[i].text
        
        masterSheet["H1"]  = datestamp
        masterSheet["H1"].fill = PatternFill("solid", fgColor="FFA7A7")

    testWB.save("FCS Bank Leadership.xlsm")
    return leadership

def cobank():
    
    browser.get("https://www.cobank.com/web/cobank/corporate/management-executive-committee")
    #tab = browser.find_elements_by_tag_name("a")
    #tab[99].click()
    #time.sleep(10)

    csuite = browser.find_element_by_id("fragment-0-olfa")

    title = csuite.find_elements_by_tag_name("p.card-text")
    name = csuite.find_elements_by_tag_name("h5.card-title")

    leadership = {}
    nameList = []
    prevLeadership = set()

    for x in range(3,20):
        while masterSheet[f"B{str(x)}"].value != None:
            prevLeadership.add(masterSheet[f"B{str(x)}"].value)
            break

    if len(name) == len(title):
        for i in range(0,len(name)):
            leadership[name[i].text] = title[i].text
            nameList.append(name[i].text)
            i += 1

        for i in range(0,len(name)):
            bankSheet["B"+ str(5+i)] = name[i].text
            bankSheet["C"+ str(5+i)] = leadership[name[i].text]
            bankSheet["D"+ str(5+i)] = datestamp
            
    setCompare = prevLeadership.symmetric_difference(nameList)

    if len(setCompare) > 0:
        bankSheet[f"B{5 + len(name)}"] = str(setCompare) 
        bankSheet[f"B{5 + len(name)}"].fill = PatternFill("solid", fgColor="FFA7A7")

        for x in range(3,20):
            while masterSheet[f"B{str(x)}"].value != None:
                masterSheet[f"B{str(x)}"].value = None
                break
        for i in range(0,len(name)):
            masterSheet["B"+ str(3+i)] = name[i].text

        masterSheet["B1"]  = datestamp
        masterSheet["B1"].fill = PatternFill("solid", fgColor="FFA7A7")

    testWB.save("FCS Bank Leadership.xlsm")
    return leadership, setCompare

def agfirst():
    
    browser.get("https://www.agfirst.com/About-Us/Leadership.aspx")    

    c_suite = browser.find_element_by_tag_name("section.content-config.leadership")

    title = c_suite.find_elements_by_tag_name("h3")
    name = c_suite.find_elements_by_tag_name("h2")

    leadership = {}
    nameList = []
    prevLeadership = set()
    
    #for item in name:
    #    nm_item = item.text        
    #    nameList.append(nm_item)    

    for x in range(3,20):
        while masterSheet[f"D{str(x)}"].value != None:
            prevLeadership.add(masterSheet[f"D{str(x)}"].value)
            break

    if len(name) == len(title):
        for i in range(0,len(name)):
            leadership[name[i].text] = title[i].text
            nameList.append(name[i].text)
            i += 1

        for i in range(0,len(name)):
            bankSheet["E"+ str(5+i)] = name[i].text
            bankSheet["F"+ str(5+i)] = leadership[name[i].text]
            bankSheet["G"+ str(5+i)] = datestamp        

    setCompare = prevLeadership.symmetric_difference(nameList)

    if len(setCompare) > 0:
        bankSheet[f"E{5 + len(name)}"] = str(setCompare) 
        bankSheet[f"E{5 + len(name)}"].fill = PatternFill("solid", fgColor="FFA7A7")

        for x in range(3,20):
            while masterSheet[f"D{str(x)}"].value != None:
                masterSheet[f"D{str(x)}"].value = None
                break

        for i in range(0,len(name)):
            masterSheet["D"+ str(3+i)] = name[i].text

        masterSheet["D1"]  = datestamp
        masterSheet["D1"].fill = PatternFill("solid", fgColor="FFA7A7")

    testWB.save("FCS Bank Leadership.xlsm")
    return leadership, setCompare

def agribank():
    
    browser.get("https://info.agribank.com/about/Pages/Executive-Officers.aspx")    
    time.sleep(3)

    csuite = browser.find_elements_by_tag_name("h6")

    title = []
    name = []

    ## Compare
    prevLeadership = set()

    for x in range(3,20):
        while masterSheet[f"F{str(x)}"].value != None:
            prevLeadership.add(masterSheet[f"F{str(x)}"].value)
            break    

    for item in csuite:
        nm_item = item.text[0:item.text.find(" - ")]
        ttl_item = item.text[item.text.find(" - ")+3:]

        title.append(ttl_item)
        name.append(nm_item)
    
    leadership = {}

    if len(name) == len(title):
        for i in range(0,len(name)):
            leadership[name[i]] = title[i]
            i += 1

        for i in range(0,len(name)):
            bankSheet["H"+ str(5+i)] = name[i]
            bankSheet["I"+ str(5+i)] = leadership[name[i]]
            bankSheet["J"+ str(5+i)] = datestamp

    setCompare = prevLeadership.symmetric_difference(name)

    if len(setCompare) > 0:
        bankSheet[f"H{5 + len(name)}"] = str(setCompare) 
        bankSheet[f"H{5 + len(name)}"].fill = PatternFill("solid", fgColor="FFA7A7")

        for x in range(3,20):
            while masterSheet[f"F{str(x)}"].value != None:
                masterSheet[f"F{str(x)}"].value = None
                break

        for i in range(0,len(name)):
            masterSheet["F"+ str(3+i)] = name[i]

        masterSheet["F1"]  = datestamp
        masterSheet["F1"].fill = PatternFill("solid", fgColor="FFA7A7")

    testWB.save("FCS Bank Leadership.xlsm")
    return leadership, setCompare

def callall():
    texasfcb()
    cobank()
    agfirst()
    agribank()

#mainSheet.title = "Main Sheet"
#c1 = mainSheet["A1"]
#mainSheet["A1"] = "Testing"

#testWB.save("Test.xlsx")

#wksht = wkbk.active
#wksht1 = wkbk.create_sheet("MySheet")

#for row in lastSheet['E']:
#    print(row.value)